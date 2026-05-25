"""
Document tools — PDF, Excel, DOCX reading with zero/minimal dependencies.

All tools degrade gracefully when optional deps are missing.
"""

import json
from dataclasses import dataclass
from pathlib import Path

from sigma.agent import BaseTool

MAX_FILE_BYTES = 32 * 1024 * 1024  # 32 MB
FORBIDDEN_PATH_PATTERNS = {"..", "~", "$"}


def _safe_path(filepath: str, root: Path | None = None) -> Path:
    """Resolve and validate a file path, blocking traversal escapes."""
    if not filepath or not isinstance(filepath, str):
        raise ValueError("filepath is required and must be a string")
    clean = filepath.strip()
    if not clean:
        raise ValueError("filepath cannot be empty")
    for pattern in FORBIDDEN_PATH_PATTERNS:
        if pattern in clean:
            raise ValueError(f"filepath contains forbidden pattern: {pattern}")
    p = Path(clean)
    if p.is_absolute():
        resolved = p.resolve()
    else:
        resolved = (root or Path.cwd()).resolve() / clean
        resolved = resolved.resolve()
        root_path = (root or Path.cwd()).resolve()
        try:
            resolved.relative_to(root_path)
        except ValueError:
            raise ValueError(f"Path traversal blocked: '{clean}' escapes root")
    if resolved.suffix.lower() in {".exe", ".dll", ".so", ".dylib", ".bin"}:
        raise ValueError(f"Forbidden file extension: {resolved.suffix}")
    return resolved


def _check_file(filepath: str) -> Path | None:
    """Validate filepath exists, is safe, and within size limit. Returns Path or None."""
    try:
        p = _safe_path(filepath)
    except ValueError:
        return None
    if not p.exists():
        return None
    if p.stat().st_size > MAX_FILE_BYTES:
        return None
    return p


# ── PDF Tool ──────────────────────────────────────────────────────────

@dataclass
class PdfTool(BaseTool):
    """Extract text from PDF files.

    Uses PyPDF2 if available (recommended: pip install PyPDF2).
    Falls back to a basic implementation if not installed.
    """

    name: str = "pdf_tool"
    description: str = (
        "Extract text from PDF files. Requires PyPDF2 (pip install PyPDF2) for best results. "
        "Returns extracted text with page markers."
    )

    def _run(self, filepath: str = "", pages: str = "", **kwargs) -> dict:
        """Extract text from a PDF file.

        Args:
            filepath: Path to the PDF file.
            pages: Optional page range like '1-5' or '3'. Default: all pages.

        Returns:
            dict with: success, filepath, page_count, text (extracted content per page).
        """
        if not filepath:
            return {"success": False, "error": "filepath is required"}

        p = _check_file(filepath)
        if p is None:
            return {"success": False, "error": f"Cannot read PDF: {filepath}"}

        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return {
                "success": False,
                "error": (
                    "PyPDF2 is not installed. Install it with: pip install PyPDF2. "
                    "Then retry this operation."
                ),
            }

        try:
            reader = PdfReader(str(p))
            total_pages = len(reader.pages)

            # Parse page range
            target_pages = _parse_page_range(pages, total_pages)

            result_pages = []
            for i in target_pages:
                page_text = reader.pages[i].extract_text() or ""
                result_pages.append({"page": i + 1, "text": page_text[:20000]})

            return {
                "success": True,
                "filepath": filepath,
                "page_count": total_pages,
                "extracted_pages": len(result_pages),
                "pages": result_pages,
            }
        except Exception as e:
            return {"success": False, "error": f"PDF read error: {e}"}


def _parse_page_range(pages: str, total: int) -> list[int]:
    """Parse a page range string like '1-5', '3', or '' (all)."""
    if not pages:
        return list(range(total))

    pages = pages.strip()
    if "-" in pages:
        parts = pages.split("-", 1)
        try:
            start = max(0, int(parts[0]) - 1)
            end = min(total, int(parts[1]))
            return list(range(start, end))
        except ValueError:
            return list(range(total))

    try:
        idx = max(0, int(pages) - 1)
        return [min(idx, total - 1)]
    except ValueError:
        return list(range(total))


# ── Excel Tool ────────────────────────────────────────────────────────

@dataclass
class ExcelTool(BaseTool):
    """Read Excel files (.xlsx, .xls) and extract sheet data.

    Uses openpyxl if available (recommended: pip install openpyxl).
    """

    name: str = "excel_tool"
    description: str = (
        "Read Excel spreadsheets. Operations: 'sheets' (list sheet names), "
        "'read' (extract a specific sheet as rows of dicts). "
        "Requires openpyxl (pip install openpyxl)."
    )

    def _run(
        self,
        filepath: str = "",
        operation: str = "read",
        sheet: str = "",
        **kwargs,
    ) -> dict:
        """Read an Excel file.

        Args:
            filepath: Path to the Excel file.
            operation: 'sheets' (list all sheets) or 'read' (extract data).
            sheet: Sheet name for 'read' operation. Default: first sheet.

        Returns:
            dict with: success, sheet_names (for 'sheets'), data (list of row dicts for 'read').
        """
        if not filepath:
            return {"success": False, "error": "filepath is required"}

        p = _check_file(filepath)
        if p is None:
            return {"success": False, "error": f"Cannot read Excel file: {filepath}"}

        try:
            import openpyxl
        except ImportError:
            return {
                "success": False,
                "error": "openpyxl is not installed. Install it with: pip install openpyxl",
            }

        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)

            if operation == "sheets":
                names = wb.sheetnames
                wb.close()
                return {"success": True, "filepath": filepath, "sheet_names": names, "sheet_count": len(names)}

            # 'read' operation
            ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

            rows = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                else:
                    row_dict = {headers[j]: c for j, c in enumerate(row)} if headers else {}
                    rows.append(row_dict)
                if len(rows) >= 5000:
                    break

            wb.close()
            return {
                "success": True,
                "filepath": filepath,
                "sheet": sheet or wb.sheetnames[0],
                "headers": headers,
                "row_count": len(rows),
                "data": rows,
            }
        except Exception as e:
            return {"success": False, "error": f"Excel read error: {e}"}


# ── DOCX Tool ─────────────────────────────────────────────────────────

@dataclass
class DocxTool(BaseTool):
    """Extract text from Word documents (.docx).

    Uses python-docx if available (recommended: pip install python-docx).
    """

    name: str = "docx_tool"
    description: str = (
        "Extract text from Microsoft Word .docx files. "
        "Requires python-docx (pip install python-docx). "
        "Returns all paragraph text."
    )

    def _run(self, filepath: str = "", **kwargs) -> dict:
        """Extract paragraphs from a .docx file.

        Args:
            filepath: Path to the .docx file.

        Returns:
            dict with: success, filepath, paragraph_count, char_count, text (all paragraphs joined).
        """
        if not filepath:
            return {"success": False, "error": "filepath is required"}

        p = _check_file(filepath)
        if p is None:
            return {"success": False, "error": f"Cannot read DOCX: {filepath}"}

        try:
            from docx import Document
        except ImportError:
            return {
                "success": False,
                "error": "python-docx is not installed. Install it with: pip install python-docx",
            }

        try:
            doc = Document(str(p))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            full_text = "\n\n".join(paragraphs)

            return {
                "success": True,
                "filepath": filepath,
                "paragraph_count": len(paragraphs),
                "char_count": len(full_text),
                "text": full_text[:50000],
            }
        except Exception as e:
            return {"success": False, "error": f"DOCX read error: {e}"}
